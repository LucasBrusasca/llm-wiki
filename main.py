import os
import random
import re
import threading
from contextlib import asynccontextmanager
from pathlib import Path
from urllib.parse import unquote

from dotenv import load_dotenv
from fastapi import BackgroundTasks, Depends, FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, Response
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
from sqlalchemy import select, text
from sqlalchemy import delete as sql_delete
from sqlalchemy.dialects.postgresql import insert as pg_insert
from sqlalchemy.ext.asyncio import AsyncSession

from database.connection import get_async_session, get_sync_session
from database.models import AuditLog, Edge, Node

load_dotenv()

BASE = Path(__file__).parent.resolve()
UPLOADS = BASE / "uploads"
UPLOADS.mkdir(exist_ok=True)
THUMBS = UPLOADS / "thumbs"
THUMBS.mkdir(exist_ok=True)


@asynccontextmanager
async def lifespan(app: FastAPI):
    from database.init_db import init_db
    await init_db()
    yield


app = FastAPI(title="Algedi", lifespan=lifespan)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# El LLM se invoca vía processor.query_llm, que respeta LLM_PROVIDER
# (ollama / gemini / anthropic). No se instancia ningún cliente acá.

# ── Ingestion state ───────────────────────────────────────────────────
_ingest = {"state": "idle", "message": "", "label": "", "progress": 0}
_ingest_lock = threading.Lock()

# ── Issue state ────────────────────────────────────────────────────────
_issue_state = {"state": "idle", "message": "", "progress": 0, "result": None}
_issue_lock = threading.Lock()


def _resolve(rel: str) -> Path:
    target = (BASE / unquote(rel)).resolve()
    if not str(target).startswith(str(BASE)):
        raise HTTPException(status_code=403, detail="Forbidden")
    if not target.exists():
        raise HTTPException(status_code=404, detail="Not found")
    return target


# ── Conversion helpers ────────────────────────────────────────────────

def node_to_dict(n) -> dict:
    emb = n.embedding
    if hasattr(emb, "tolist"):
        emb = emb.tolist()
    if emb:
        # 4 decimales bastan para el coseno del frontend (descubrimientos/centroide)
        # y achican ~60% el peso del payload de /api/graph.
        emb = [round(float(x), 4) for x in emb]
    return {
        "id": n.id,
        "label": n.label,
        "type": n.type,
        "desc": n.desc,
        "fragmento": n.fragmento,
        "conceptos": n.conceptos or [],
        "embedding": emb,
        "x3d": n.x3d,
        "y3d": n.y3d,
        "z3d": n.z3d,
        "x_pca": n.x_pca,
        "y_pca": n.y_pca,
        "z_pca": n.z_pca,
        "cluster": n.cluster if n.cluster is not None else -1,
        "dominio": n.dominio or "personal",
        "fuente": n.fuente,
        "fuente_url": n.fuente_url,
        "fuente_path": n.fuente_path,
        "fuente_label": n.fuente_label,
        "autor": n.autor,
        "fecha_doc": n.fecha_doc,
        "tema": n.tema,
        "flujograma": n.flujograma,
        "synthesis": n.synthesis,
        "is_centroid": n.is_centroid or False,
        "is_issue": n.is_issue or False,
        "tags": n.tags or [],
        # rich_html NO viaja acá: puede pesar cientos de KB por nodo y el frontend
        # lo pide on-demand vía /api/node/{id}/rich-preview. Sacarlo aliviana /api/graph.
        "created_at": n.created_at.isoformat() if n.created_at else None,
    }


def edge_to_dict(e) -> dict:
    return {
        "source": e.source,
        "target": e.target,
        "score": e.score,
        "shared_concepts": e.shared_concepts or [],
        "label": e.label,
        "description": e.description,
    }


# ── Helpers ───────────────────────────────────────────────────────────

async def _nodos_relevantes(pregunta: str, max_n: int = 5,
                             db: AsyncSession = None) -> list[str]:
    try:
        from processor import get_embed_model
        model = get_embed_model()
        vec = model.encode([pregunta], show_progress_bar=False)[0].tolist()
        result = await db.execute(
            text("""
                SELECT id
                FROM nodes
                WHERE NOT COALESCE(is_centroid, false)
                  AND embedding IS NOT NULL
                ORDER BY embedding <=> CAST(:vec AS vector)
                LIMIT :lim
            """),
            {"vec": str(vec), "lim": max_n}
        )
        return [row.id for row in result]
    except Exception as e:
        print(f"Vector search fallback (keyword): {e}")
        stops = {"el","la","los","las","un","una","de","en","que","es",
                 "y","a","con","por","para","del","al","se","no","lo","su","sus"}
        words = [w for w in pregunta.lower().split()
                 if len(w) > 2 and w not in stops]
        if not words:
            return []
        rows = (await db.execute(
            select(Node.id, Node.label, Node.desc)
            .where(Node.is_centroid == False)
        )).all()
        scored = []
        for row in rows:
            txt = f"{row.label or ''} {row.desc or ''}".lower()
            s = sum(1 for w in words if w in txt)
            if s > 0:
                scored.append((s, row.id))
        scored.sort(reverse=True)
        return [nid for _, nid in scored[:max_n]]


async def _nodos_relevantes_scored(pregunta: str, max_n: int = 6,
                                   db: AsyncSession = None) -> list[dict]:
    """Como _nodos_relevantes pero devuelve también el contenido y la SIMILITUD coseno
    (sim = 1 - distancia). Lo usa el agente para fundamentar la respuesta y para el VETO."""
    try:
        from processor import get_embed_model
        model = get_embed_model()
        vec = model.encode([pregunta], show_progress_bar=False)[0].tolist()
        result = await db.execute(
            text("""
                SELECT id, label, "desc", fragmento,
                       (embedding <=> CAST(:vec AS vector)) AS dist
                FROM nodes
                WHERE NOT COALESCE(is_centroid, false)
                  AND NOT COALESCE(is_issue, false)
                  AND embedding IS NOT NULL
                ORDER BY dist
                LIMIT :lim
            """),
            {"vec": str(vec), "lim": max_n}
        )
        out = []
        for r in result:
            out.append({
                "id": r.id, "label": r.label, "desc": r.desc, "fragmento": r.fragmento,
                "sim": max(0.0, 1.0 - float(r.dist)),
            })
        return out
    except Exception as e:
        print(f"scored retrieval failed: {e}")
        return []


def _save_node_sync(nodo_data: dict):
    """Guarda o actualiza un nodo en PostgreSQL y recalcula relaciones. Síncrona — segura para threads."""
    from database.models import Node as NodeModel, Edge as EdgeModel
    from sqlalchemy import delete as sync_delete
    from processor import _auto_relaciones

    campos_validos = {c.key for c in NodeModel.__table__.columns}
    datos = {k: v for k, v in nodo_data.items() if k in campos_validos}

    with get_sync_session() as session:
        session.execute(
            pg_insert(NodeModel).values(**datos)
            .on_conflict_do_update(index_elements=["id"], set_=datos)
        )
        session.flush()

        # Los issues/procesos NO participan de las relaciones del grafo de conocimiento:
        # son entidades de otro plano (módulo Issue/Procesos) cuyo fundamento se calcula
        # on-the-fly contra el grafo. Mezclarlos ensuciaba el espacio de documentos.
        all_nodes = session.query(NodeModel).filter(
            NodeModel.is_centroid == False,
            NodeModel.is_issue == False,
        ).all()

        nodes_dicts = []
        for n in all_nodes:
            emb = n.embedding
            if hasattr(emb, "tolist"):
                emb = emb.tolist()
            nodes_dicts.append({
                "id": n.id,
                "label": n.label,
                "conceptos": n.conceptos or [],
                "embedding": emb,
                "is_centroid": False,
            })

        new_rels = _auto_relaciones(nodes_dicts)

        session.execute(sync_delete(EdgeModel))
        for r in new_rels:
            session.add(EdgeModel(
                source=r["source"],
                target=r["target"],
                score=r.get("score"),
                shared_concepts=r.get("shared_concepts", []),
                label=r.get("label"),
                description=r.get("description"),
            ))
        session.commit()


def _recompute_edges_background():
    from processor import _auto_relaciones
    from database.models import Node as NodeModel, Edge as EdgeModel
    from sqlalchemy import delete as sync_delete
    with get_sync_session() as session:
        nodes = session.query(NodeModel).filter(
            NodeModel.is_centroid == False,
            NodeModel.is_issue == False,   # issues fuera del grafo de conocimiento
        ).all()
        nodes_dicts = []
        for n in nodes:
            emb = n.embedding
            if hasattr(emb, "tolist"):
                emb = emb.tolist()
            nodes_dicts.append({
                "id": n.id, "label": n.label,
                "conceptos": n.conceptos or [],
                "embedding": emb, "is_centroid": False,
            })
        new_rels = _auto_relaciones(nodes_dicts)
        session.execute(sync_delete(EdgeModel))
        for r in new_rels:
            session.add(EdgeModel(**r))
        session.commit()


# ── Graph data ────────────────────────────────────────────────────────

@app.get("/api/graph")
async def get_graph(db: AsyncSession = Depends(get_async_session)):
    nodes_rows = (await db.execute(select(Node))).scalars().all()
    edges_rows = (await db.execute(select(Edge))).scalars().all()

    rng = random.Random(42)
    nodes_list = []
    for n in nodes_rows:
        nd = node_to_dict(n)
        if nd["x3d"] is None:
            nd["x3d"] = rng.uniform(-1, 1)
            nd["y3d"] = rng.uniform(-1, 1)
            nd["z3d"] = rng.uniform(-1, 1)
        nodes_list.append(nd)

    return {
        "nodos": nodes_list,
        "relaciones": [edge_to_dict(e) for e in edges_rows],
    }


# ── File serving ──────────────────────────────────────────────────────

@app.get("/thumbnail")
def get_thumbnail(p: str):
    target = _resolve(p)
    try:
        import fitz
        doc = fitz.open(str(target))
        pix = doc[0].get_pixmap(matrix=fitz.Matrix(1.2, 1.2), alpha=False)
        data = pix.tobytes("png")
        doc.close()
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    return Response(
        content=data,
        media_type="image/png",
        headers={"Cache-Control": "public, max-age=86400"},
    )


@app.get("/doc")
def get_doc(p: str):
    target = _resolve(p)
    ext = target.suffix.lower().lstrip(".")
    media = {
        "pdf":  "application/pdf",
        "mp3":  "audio/mpeg",
        "wav":  "audio/wav",
        "ogg":  "audio/ogg",
        "html": "text/html",
        "htm":  "text/html",
        "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "xls":  "application/vnd.ms-excel",
    }.get(ext, "application/octet-stream")
    disposition = "inline" if ext in ("pdf", "html", "htm") else "attachment"
    return Response(
        content=target.read_bytes(),
        media_type=media,
        headers={"Content-Disposition": f'{disposition}; filename="{target.name}"'},
    )


@app.get("/excel-preview")
def excel_preview(p: str):
    target = _resolve(p)
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(target), read_only=True, data_only=True)
        ws = wb.active
        rows = []
        for i, row in enumerate(ws.iter_rows(max_row=5, max_col=5, values_only=True)):
            rows.append([str(c) if c is not None else "" for c in row])
            if i >= 4:
                break
        wb.close()
        return {"rows": rows}
    except ImportError:
        raise HTTPException(status_code=501, detail="openpyxl not installed")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


def _resolve_file(p: str) -> Path:
    """Resuelve la ruta de un archivo ingerido, tolerando rutas absolutas del
    contenedor, relativas o solo el nombre dentro de UPLOADS. Siempre confinado a BASE."""
    raw = unquote(p or "").replace("\\", "/")
    candidates = []
    pp = Path(raw)
    candidates.append(pp)
    if not pp.is_absolute():
        candidates.append(BASE / raw)
    candidates.append(UPLOADS / Path(raw).name)  # fallback: por nombre
    for c in candidates:
        try:
            r = c.resolve()
        except Exception:
            continue
        if str(r).startswith(str(BASE)) and r.exists() and r.is_file():
            return r
    raise HTTPException(status_code=404, detail="Archivo no encontrado")


@app.get("/files/{node_id}")
async def get_file(node_id: str, db: AsyncSession = Depends(get_async_session)):
    """Sirve el archivo original asociado a un nodo (por id). Inline para que
    PDF/imagen/audio/video se previsualicen embebidos. Soporta Range (seek)."""
    node = (await db.execute(select(Node).where(Node.id == node_id))).scalar_one_or_none()
    if node is None:
        raise HTTPException(status_code=404, detail="Nodo no encontrado")
    if not node.fuente_path:
        raise HTTPException(status_code=404, detail="El nodo no tiene archivo")
    target = _resolve_file(node.fuente_path)
    import mimetypes
    media, _ = mimetypes.guess_type(target.name)
    return FileResponse(
        str(target),
        media_type=media or "application/octet-stream",
        headers={"Content-Disposition": f'inline; filename="{target.name}"'},
    )


def _yt_id(url: str):
    import re as _re
    m = _re.search(r"(?:youtu\.be/|v=|embed/)([A-Za-z0-9_-]{11})", url or "")
    return m.group(1) if m else None


def _render_thumb(doc) -> bytes:
    """Renderiza la primera página/imagen de un documento fitz a PNG ~128px."""
    import fitz
    page = doc[0]
    rect = page.rect
    big = max(rect.width, rect.height) or 1
    zoom = min(128.0 / big, 2.0)  # no agrandar de más
    pix = page.get_pixmap(matrix=fitz.Matrix(zoom, zoom), alpha=False)
    return pix.tobytes("png")


def _text_thumb(texto: str) -> bytes | None:
    """Miniatura a partir de TEXTO (para formatos que fitz no abre: Word/PPT/Excel).
    Renderiza el contenido como página de texto → muestra el inicio del documento."""
    import fitz
    snippet = (texto or "").strip()[:1500]
    if not snippet:
        return None
    try:
        doc = fitz.open(stream=snippet.encode("utf-8"), filetype="txt")
        data = _render_thumb(doc); doc.close()
        return data
    except Exception:
        return None


def _generate_thumb(node) -> bytes | None:
    """Miniatura (PNG ~128px) según el tipo de nodo. None si no aplica."""
    import fitz
    fuente = (node.fuente or "").lower()

    # YouTube: usar el thumbnail de la URL (descargado server-side → mismo origen, CORS-clean)
    if fuente == "youtube" and node.fuente_url:
        vid = _yt_id(node.fuente_url)
        if not vid:
            return None
        import httpx
        for q in ("mqdefault", "hqdefault", "default"):
            try:
                r = httpx.get(f"https://img.youtube.com/vi/{vid}/{q}.jpg", timeout=10.0)
                if r.status_code == 200 and len(r.content) > 1000:
                    doc = fitz.open(stream=r.content, filetype="jpg")
                    data = _render_thumb(doc); doc.close()
                    return data
            except Exception:
                continue
        return None

    # Archivos locales (PDF / imagen). fitz abre tanto PDF como imágenes.
    if node.fuente_path:
        try:
            target = _resolve_file(node.fuente_path)
        except HTTPException:
            return None
        ext = target.suffix.lower().lstrip(".")
        if ext == "pdf" or fuente in ("pdf", "tesis") or ext in (
            "png", "jpg", "jpeg", "gif", "webp", "bmp", "tif", "tiff"
        ):
            try:
                doc = fitz.open(str(target))
                data = _render_thumb(doc); doc.close()
                return data
            except Exception:
                return None
        # HTML: fitz lo renderiza como documento → miniatura de la primera parte.
        if ext in ("html", "htm") or fuente == "html":
            try:
                doc = fitz.open(str(target), filetype="html")
                data = _render_thumb(doc); doc.close()
                return data
            except Exception:
                return None
        # Texto plano / Markdown: fitz lo pagina como texto.
        if ext in ("txt", "md", "markdown"):
            try:
                return _text_thumb(target.read_text(encoding="utf-8", errors="ignore"))
            except Exception:
                return None
        # Word / PowerPoint: fitz NO los abre → miniatura del texto extraído.
        if ext == "docx" or fuente == "word":
            from processor import _extraer_texto_office
            return _text_thumb(_extraer_texto_office(str(target), ["word/document.xml"]))
        if ext in ("pptx", "pptm") or fuente == "ppt":
            from processor import _extraer_texto_office
            return _text_thumb(_extraer_texto_office(str(target), ["ppt/slides/"]))
        # Excel: primeras filas como texto.
        if ext in ("xlsx", "xls") or fuente == "excel":
            try:
                import openpyxl
                wb = openpyxl.load_workbook(str(target), read_only=True, data_only=True)
                ws = wb.active
                rows = []
                for i, row in enumerate(ws.iter_rows(max_row=20, values_only=True)):
                    txt = " | ".join(str(c) for c in row if c is not None)
                    if txt.strip():
                        rows.append(txt)
                    if i >= 19:
                        break
                wb.close()
                return _text_thumb("\n".join(rows))
            except Exception:
                return None
    return None


def _render_html_playwright(path: str) -> bytes | None:
    """Miniatura de un HTML renderizado CON estilos (Chromium headless): captura del
    inicio del documento → reducida. Pixel-perfect. SÍNCRONA: llamar en un thread
    (la sync API de Playwright no corre dentro del event loop de asyncio)."""
    try:
        from playwright.sync_api import sync_playwright
        with sync_playwright() as p:
            browser = p.chromium.launch(args=["--no-sandbox"])
            page = browser.new_page(viewport={"width": 900, "height": 650}, device_scale_factor=1)
            page.goto(f"file://{path}", wait_until="networkidle", timeout=15000)
            png = page.screenshot(clip={"x": 0, "y": 0, "width": 900, "height": 650})
            browser.close()
        import fitz
        doc = fitz.open(stream=png, filetype="png")
        data = _render_thumb(doc); doc.close()
        return data
    except Exception as e:
        print(f"playwright html thumb failed: {e}")
        return None


@app.get("/thumb/{node_id}")
async def get_thumb(node_id: str, db: AsyncSession = Depends(get_async_session)):
    """Miniatura del archivo del nodo (cara del nodo en el grafo). 404 → el front
    cae al punto cian. Cachea en disco para no re-renderizar (perf en GPU integrada)."""
    import hashlib
    h = hashlib.md5(node_id.encode("utf-8")).hexdigest()
    cache = THUMBS / f"{h}.png"
    if cache.exists():
        return FileResponse(str(cache), media_type="image/png",
                            headers={"Cache-Control": "public, max-age=86400"})
    node = (await db.execute(select(Node).where(Node.id == node_id))).scalar_one_or_none()
    if node is None:
        raise HTTPException(status_code=404, detail="Nodo no encontrado")
    data = None
    # HTML → render CON estilos (Chromium headless), en un thread. Fallback a fitz si falla.
    fuente = (node.fuente or "").lower()
    fp = (node.fuente_path or "").lower()
    if node.fuente_path and (fuente == "html" or fp.endswith((".html", ".htm"))):
        try:
            import asyncio
            target = _resolve_file(node.fuente_path)
            data = await asyncio.to_thread(_render_html_playwright, str(target))
        except Exception:
            data = None
    if not data:
        try:
            data = _generate_thumb(node)
        except Exception:
            data = None
    if not data:
        raise HTTPException(status_code=404, detail="Sin miniatura")
    try:
        cache.write_bytes(data)
    except Exception:
        pass
    return Response(content=data, media_type="image/png",
                    headers={"Cache-Control": "public, max-age=86400"})


# ── Reset / clear ─────────────────────────────────────────────────────

@app.post("/api/reset")
async def reset_graph(db: AsyncSession = Depends(get_async_session)):
    await db.execute(sql_delete(Edge))
    await db.execute(sql_delete(Node))
    await db.commit()
    umap_pkl = BASE / "umap_model.pkl"
    if umap_pkl.exists():
        umap_pkl.unlink()
    return {"ok": True}


# ── Recompute relations ───────────────────────────────────────────────

@app.post("/api/recompute-relations")
async def recompute_relations(db: AsyncSession = Depends(get_async_session)):
    from processor import _auto_relaciones
    nodes_rows = (await db.execute(
        select(Node).where(Node.is_centroid == False)
    )).scalars().all()

    nodes_dicts = []
    for n in nodes_rows:
        emb = n.embedding
        if hasattr(emb, "tolist"):
            emb = emb.tolist()
        nodes_dicts.append({
            "id": n.id,
            "label": n.label,
            "conceptos": n.conceptos or [],
            "embedding": emb,
            "is_centroid": False,
        })

    new_rels = _auto_relaciones(nodes_dicts)

    await db.execute(sql_delete(Edge))
    for r in new_rels:
        db.add(Edge(
            source=r["source"], target=r["target"],
            score=r.get("score"),
            shared_concepts=r.get("shared_concepts", []),
            label=r.get("label"),
            description=r.get("description"),
        ))
    await db.commit()
    return {"ok": True, "relaciones": len(new_rels)}


@app.post("/api/recompute-layout")
async def recompute_layout():
    """Re-corre el clustering (UMAP intermedia + HDBSCAN) sobre todos los nodos y
    persiste posiciones/clusters. Sirve para aplicar ajustes de clustering al grafo
    existente, sin necesidad de re-subir documentos."""
    import asyncio
    import embeddings_engine
    await asyncio.to_thread(embeddings_engine.main)
    return {"ok": True}


@app.post("/api/taxonomy")
async def taxonomy(apply: bool = False, db: AsyncSession = Depends(get_async_session)):
    """Propone (dry-run) o aplica una taxonomía de TEMAS legible, asignada por LLM.
    Agrupa el grafo como lo haría una persona (por tema), no por densidad de embeddings.
    Sin apply → solo propone, no escribe. apply=true → persiste node.tema.
    Los embeddings/UMAP siguen rigiendo la POSICIÓN 3D; esto rige el GRUPO (color/etiqueta)."""
    rows = (await db.execute(
        select(Node).where(Node.is_centroid == False, Node.is_issue == False)
    )).scalars().all()
    docs = [{"id": n.id, "label": n.label or "", "conceptos": (n.conceptos or [])[:8]} for n in rows]
    if len(docs) < 3:
        return {"ok": False, "error": f"Solo {len(docs)} nodos; muy pocos para una taxonomía."}

    listado = "\n".join(
        f'- [{d["id"]}] {d["label"]} | conceptos: {", ".join(d["conceptos"]) or "(ninguno)"}'
        for d in docs
    )
    system = (
        "Sos un bibliotecario experto que organiza un grafo de conocimiento personal sobre IA, "
        "machine learning, estadística, RAG y temas afines. Agrupás documentos por TEMA, "
        "como lo haría una persona, no por palabras sueltas."
    )
    user = f"""Tengo {len(docs)} documentos. Para cada uno te doy id, título y conceptos clave.

1) Proponé una taxonomía de entre 6 y 10 TEMAS con nombres cortos y legibles en español
   (ej: "Estadística", "LLMs", "RAG / Retrieval", "Diffusion y Visión",
   "Redes neuronales / Deep learning", "ML clásico", "Agentes", "IA y sociedad").
   Los temas deben reflejar los temas REALES de ESTOS documentos, no una lista genérica.
2) Asigná CADA documento a exactamente UN tema de tu taxonomía.
   Si un documento no encaja claramente en ninguno, asignalo a "Sin clasificar" — no lo fuerces.

Documentos:
{listado}

Respondé SOLO con JSON válido, sin backticks:
{{
  "temas": ["Tema 1", "Tema 2", "..."],
  "asignaciones": [{{"id": "<id exacto>", "tema": "<uno de temas, o 'Sin clasificar'>"}}]
}}
Cada id debe aparecer exactamente una vez. No inventes ids."""

    from processor import query_llm, parsear_json
    import asyncio
    from collections import Counter
    raw = await asyncio.to_thread(query_llm, [{"role": "user", "content": user}], system)
    try:
        data = parsear_json(raw)
    except Exception:
        return {"ok": False, "error": "El LLM no devolvió JSON válido.", "raw": raw[:600]}

    temas = data.get("temas") or []
    asign = {}
    for a in data.get("asignaciones", []):
        if isinstance(a, dict) and a.get("id"):
            asign[a["id"]] = (a.get("tema") or "Sin clasificar").strip()

    conteo = Counter(asign.get(d["id"], "Sin clasificar") for d in docs)

    if apply:
        for n in rows:
            n.tema = asign.get(n.id) or "Sin clasificar"
        await db.commit()

    return {
        "ok": True,
        "applied": apply,
        "temas": temas,
        "conteo": dict(conteo),
        "asignaciones": [
            {"id": d["id"], "label": d["label"], "tema": asign.get(d["id"], "Sin clasificar")}
            for d in docs
        ],
    }


# ── Search ────────────────────────────────────────────────────────────

@app.get("/api/search")
async def semantic_search(
    q: str,
    max_n: int = 10,
    db: AsyncSession = Depends(get_async_session),
):
    if not q or len(q.strip()) < 2:
        return {"ids": []}
    ids = await _nodos_relevantes(q.strip(), max_n=max_n, db=db)
    return {"ids": ids}


# ── Node operations ───────────────────────────────────────────────────

@app.delete("/api/node/{node_id}")
async def delete_node(
    node_id: str,
    background_tasks: BackgroundTasks,
    db: AsyncSession = Depends(get_async_session),
):
    result = await db.execute(
        sql_delete(Node).where(Node.id == node_id)
    )
    if result.rowcount == 0:
        raise HTTPException(404, f"Nodo {node_id} no encontrado")
    await db.commit()
    background_tasks.add_task(_recompute_edges_background)
    return {"ok": True}


@app.put("/api/node/{node_id}/rename")
async def rename_node(
    node_id: str,
    request: Request,
    db: AsyncSession = Depends(get_async_session),
):
    body = await request.json()
    new_label = (body.get("label") or "").strip()
    if not new_label:
        raise HTTPException(400, "Label vacío")
    result = await db.execute(select(Node).where(Node.id == node_id))
    node = result.scalar_one_or_none()
    if not node:
        raise HTTPException(404, f"Nodo {node_id} no encontrado")
    node.label = new_label
    await db.commit()
    return {"ok": True}


class TagsUpdate(BaseModel):
    tags: list[str]


@app.post("/api/node/{node_id}/tags")
async def update_node_tags(
    node_id: str,
    payload: TagsUpdate,
    db: AsyncSession = Depends(get_async_session),
):
    result = await db.execute(select(Node).where(Node.id == node_id))
    node = result.scalar_one_or_none()
    if not node:
        raise HTTPException(404, "Nodo no encontrado")
    node.tags = payload.tags
    await db.commit()
    return {"ok": True, "tags": payload.tags}


@app.get("/api/node/{node_id}/report")
async def get_node_report(
    node_id: str,
    db: AsyncSession = Depends(get_async_session),
):
    node = (await db.execute(select(Node).where(Node.id == node_id))).scalar_one_or_none()
    if not node:
        raise HTTPException(404, f"Nodo con ID {node_id} no encontrado")

    edges_rows = (await db.execute(select(Edge))).scalars().all()
    connected_ids = set()
    for e in edges_rows:
        if e.source == node_id:
            connected_ids.add(e.target)
        elif e.target == node_id:
            connected_ids.add(e.source)

    connected_nodes = []
    if connected_ids:
        rows = (await db.execute(
            select(Node).where(Node.id.in_(connected_ids))
        )).scalars().all()
        connected_nodes = [node_to_dict(n) for n in rows]

    nd = node_to_dict(node)

    md = "# Reporte de Conexiones Semánticas\n\n"
    md += f"## Nodo Principal: {nd.get('label', 'Sin título')}\n"
    md += f"- **ID**: `{nd['id']}`\n"
    md += f"- **Fuente**: {(nd.get('fuente') or 'concepto').upper()}\n"
    if nd.get("cluster") is not None:
        md += f"- **Cluster**: {nd['cluster']}\n"
    md += "\n"
    md += "### Descripción\n"
    md += f"> {nd.get('desc', 'Sin descripción.')}\n\n"
    if nd.get("fragmento"):
        md += "### Fragmento Extraído\n"
        md += f"```text\n{nd['fragmento']}\n```\n\n"
    if nd.get("conceptos"):
        md += "### Conceptos Clave\n"
        md += ", ".join([f"`{c}`" for c in nd["conceptos"]]) + "\n\n"
    md += "---\n\n"
    md += f"## Conexiones y Relaciones ({len(connected_nodes)})\n\n"

    if not connected_nodes:
        md += "*Este nodo no tiene conexiones directas con otros nodos.*\n"
    else:
        for idx, conn in enumerate(connected_nodes, 1):
            md += f"### {idx}. {conn.get('label', 'Sin título')} (`{conn['id']}`)\n"
            md += f"- **Fuente**: {(conn.get('fuente') or 'concepto').upper()}\n"
            if conn.get("cluster") is not None:
                md += f"- **Cluster**: {conn['cluster']}\n"
            md += f"- **Descripción**: {conn.get('desc', 'Sin descripción.')}\n"
            if conn.get("fragmento"):
                md += f"- **Fragmento**: *\"{conn['fragmento']}\"*\n"
            shared = set(nd.get("conceptos", [])) & set(conn.get("conceptos", []))
            if shared:
                md += "- **Conceptos Compartidos**: " + ", ".join([f"`{c}`" for c in shared]) + "\n"
            try:
                from processor import calcular_similitud_coseno
                sim = calcular_similitud_coseno(nd.get("embedding"), conn.get("embedding"))
                if sim and sim > 0:
                    md += f"- **Similitud Semántica**: `{sim:.2%}`\n"
            except Exception:
                pass
            md += "\n"

    filename = f"reporte-{node_id}.md"
    return Response(
        content=md,
        media_type="text/markdown",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/api/node/{node_id}/rich-preview")
async def get_rich_preview(
    node_id: str,
    db: AsyncSession = Depends(get_async_session),
):
    node = (await db.execute(select(Node).where(Node.id == node_id))).scalar_one_or_none()
    if not node:
        raise HTTPException(404, f"Nodo {node_id} no encontrado")

    if node.rich_html:
        return Response(content=node.rich_html, media_type="text/html")

    nodes_rows = (await db.execute(select(Node))).scalars().all()
    edges_rows = (await db.execute(select(Edge))).scalars().all()
    nodos = [node_to_dict(n) for n in nodes_rows]
    relaciones = [edge_to_dict(e) for e in edges_rows]
    nd = node_to_dict(node)

    from processor import generar_rich_html
    try:
        html_content = generar_rich_html(nd, nodos, relaciones)
        node.rich_html = html_content
        await db.commit()
    except Exception as e:
        raise HTTPException(500, f"Error al generar el apunte: {e}")

    return Response(content=html_content, media_type="text/html")


# ── Agent ─────────────────────────────────────────────────────────────

@app.post("/api/agent")
async def agent_endpoint(
    request: Request,
    db: AsyncSession = Depends(get_async_session),
):
    body = await request.json()
    system = body.get("system", "Sos el agente de Algedi.")
    messages = body.get("messages", [])
    ultima = messages[-1]["content"] if messages else ""

    # Recuperar del grafo CON score → fundamentar la respuesta (RAG real) + VETO epistémico.
    scored = await _nodos_relevantes_scored(ultima, max_n=6, db=db) if ultima else []
    node_ids = [s["id"] for s in scored]
    max_sim = scored[0]["sim"] if scored else 0.0

    # VETO: si la mejor coincidencia es muy débil, el agente se abstiene (no inventa).
    if ultima and max_sim < AGENT_VETO_UMBRAL:
        reply = (
            "🚫 **No tengo suficiente respaldo en tu grafo para responder esto con confianza.**\n\n"
            f"La mejor coincidencia es de apenas **{int(max_sim * 100)}%** de similitud — por debajo del umbral. "
            "Prefiero callarme antes que inventarte algo que suene bien pero no esté fundamentado en tus documentos.\n\n"
            "Probá cargar material sobre el tema, o reformular hacia algo que sí esté en tu corpus."
        )
        db.add(AuditLog(query=ultima, agent_mode="VETO", node_ids_consulted=node_ids, response=reply[:2000]))
        await db.commit()
        return {"reply": reply, "nodos_relevantes": node_ids, "veto": True, "max_sim": round(max_sim, 2)}

    # Grounding: inyectar el conocimiento recuperado CON su % de afinidad. El LLM usa esos
    # números para auto-vetar con criterio (decir "no me alcanza" si la afinidad es baja).
    if scored:
        contexto = "\n".join(
            f"- [{int(s['sim'] * 100)}% afinidad] {s['label']}: {(s.get('desc') or s.get('fragmento') or '').strip()[:240]}"
            for s in scored
        )
        system = (
            f"{system}\n\nCONOCIMIENTO RELEVANTE DE TU GRAFO (con % de afinidad a la pregunta). "
            f"Fundamentá la respuesta SÓLO en esto. Si las afinidades son bajas (≲45%) o el "
            f"contenido no cubre la pregunta, DECILO con honestidad y NO inventes:\n{contexto}"
        )
    try:
        from processor import query_llm
        reply = query_llm(messages, system)
    except Exception as e:
        reply = f"Error: {e}"

    db.add(AuditLog(
        query=ultima,
        agent_mode=system[:50],
        node_ids_consulted=node_ids,
        response=reply[:2000],
    ))
    await db.commit()

    return {"reply": reply, "nodos_relevantes": node_ids, "veto": False, "max_sim": round(max_sim, 2)}


@app.post("/api/synthesize")
async def synthesize_endpoint(request: Request, db: AsyncSession = Depends(get_async_session)):
    """Chat del módulo Issue: responde una consulta sobre un proceso/problema, fundándola
    en el grafo. Incluye SIEMPRE el contenido de los node_ids en foco (el issue y su etapa)
    + recupera conocimiento relevante por similitud. Honesto: si el grafo no cubre la
    pregunta, lo dice en vez de inventar. Devuelve markdown en 'result'."""
    import asyncio
    body = await request.json()
    query = (body.get("query") or "").strip()
    node_ids = body.get("node_ids") or []
    if not query:
        return {"result": "Escribí una consulta."}

    # Contenido explícito de los nodos en foco (el issue / la etapa seleccionada).
    contexto_focos = ""
    if node_ids:
        rows = (await db.execute(select(Node).where(Node.id.in_(node_ids)))).scalars().all()
        if rows:
            contexto_focos = "\n".join(
                f"- {n.label}: {(n.desc or n.fragmento or '').strip()[:400]}" for n in rows
            )

    # Recuperación semántica del resto del grafo para fundamentar (excluye los focos).
    scored = await _nodos_relevantes_scored(query, max_n=6, db=db)
    foco_set = set(node_ids)
    scored = [s for s in scored if s["id"] not in foco_set]
    max_sim = scored[0]["sim"] if scored else 0.0
    contexto_grafo = "\n".join(
        f"- [{int(s['sim'] * 100)}% afinidad] {s['label']}: "
        f"{(s.get('desc') or s.get('fragmento') or '').strip()[:240]}"
        for s in scored
    ) or "(sin coincidencias relevantes en el grafo)"

    system = (
        "Sos el agente de Algedi ayudando a diagnosticar y mejorar un proceso o problema. "
        "Respondé en español, con markdown, conciso y accionable. "
        "Fundamentá SÓLO en el CONTEXTO EN FOCO y el CONOCIMIENTO DEL GRAFO de abajo. "
        "Si las afinidades del grafo son bajas (≲45%) o no cubren la pregunta, DECILO con "
        "honestidad y ofrecé lo que sí puedas desde el contexto en foco, sin inventar datos.\n\n"
        f"CONTEXTO EN FOCO (el proceso/etapa):\n{contexto_focos or '(sin contexto específico)'}\n\n"
        f"CONOCIMIENTO RELEVANTE DEL GRAFO (con % de afinidad):\n{contexto_grafo}"
    )

    try:
        from processor import query_llm
        reply = await asyncio.to_thread(query_llm, [{"role": "user", "content": query}], system)
    except Exception as e:
        reply = f"Error al consultar el agente: {e}"

    consultados = list(node_ids) + [s["id"] for s in scored]
    db.add(AuditLog(query=query[:2000], agent_mode="ISSUE_CHAT",
                    node_ids_consulted=consultados, response=reply[:2000]))
    await db.commit()

    return {"result": reply, "nodos_relevantes": [s["id"] for s in scored], "max_sim": round(max_sim, 2)}


# ── Procesos (inteligencia de procesos, fundada en el grafo) ──────────

@app.post("/api/process")
async def generar_proceso(
    request: Request,
    db: AsyncSession = Depends(get_async_session),
):
    """A partir de la descripción de un proceso: lo estructura como flujograma Y lo analiza
    FUNDADO en el grafo de conocimiento (sugerencias/riesgos citando documentos, y qué medir
    y por qué — sin inventar números). MVP: genera y devuelve, no persiste todavía."""
    body = await request.json()
    descripcion = (body.get("descripcion") or "").strip()
    if not descripcion:
        raise HTTPException(400, "Falta la descripción del proceso")

    # Grounding: recuperar conocimiento relevante del grafo.
    scored = await _nodos_relevantes_scored(descripcion, max_n=6, db=db)
    contexto = "\n".join(
        f"- [{int(s['sim'] * 100)}% afinidad] {s['label']}: {(s.get('desc') or '').strip()[:200]}"
        for s in scored
    ) or "(sin conocimiento relevante en el grafo)"
    nodos_fundamento = [s["id"] for s in scored]
    afinidad_max = round(scored[0]["sim"], 2) if scored else 0.0

    prompt = (
        "Sos un analista de procesos. A partir de la descripción, generá la estructura del "
        "proceso como flujograma Y un análisis fundado en el conocimiento disponible.\n\n"
        f"DESCRIPCIÓN DEL PROCESO:\n{descripcion}\n\n"
        "CONOCIMIENTO RELEVANTE DEL GRAFO (con % de afinidad). Fundá las sugerencias/riesgos "
        "en esto; si la afinidad es baja o no aplica, DECILO y no inventes respaldo:\n"
        f"{contexto}\n\n"
        "Devolvé SOLO un JSON con esta estructura EXACTA:\n"
        '{\n'
        '  "titulo": "nombre corto del proceso",\n'
        '  "pasos": [{"id": "p1", "label": "texto del paso (<=8 palabras)", "tipo": "inicio|paso|decision|fin"}],\n'
        '  "conexiones": [{"desde": "p1", "hasta": "p2", "condicion": "opcional (sí/no en decisiones)"}],\n'
        '  "sugerencias": [{"tipo": "mejora|riesgo|automatizacion", "texto": "...", "fundamento": "documento que lo respalda, o \'sin respaldo en el grafo\'"}],\n'
        '  "indicadores": [{"que": "qué medir", "como": "cómo medirlo", "porque": "por qué importa"}]\n'
        '}\n\n'
        "Reglas: el primer paso es 'inicio' y el último 'fin'; los 'decision' pueden tener 2+ "
        "salidas con condicion; en 'indicadores' NO inventes números, proponé QUÉ medir y POR QUÉ; "
        "español; SOLO el JSON, sin texto alrededor."
    )

    try:
        from processor import query_llm, parsear_json
        raw = query_llm(
            [{"role": "user", "content": prompt}],
            system="Sos un experto en análisis y optimización de procesos. Respondés SOLO con JSON válido.",
        )
        data = parsear_json(raw.strip())
    except Exception as e:
        raise HTTPException(500, f"No se pudo generar el proceso: {e}")

    data.setdefault("pasos", [])
    data.setdefault("conexiones", [])
    data.setdefault("sugerencias", [])
    data.setdefault("indicadores", [])
    data["nodos_fundamento"] = nodos_fundamento
    data["afinidad_max"] = afinidad_max
    return data


# ── Ingestion ─────────────────────────────────────────────────────────

def _set_progress(pct: int, msg: str):
    with _ingest_lock:
        # No pisar un estado terminal (error/done): el ticker corre en otro thread
        # y antes ocultaba el mensaje real del error.
        if _ingest.get("state") not in ("error", "done"):
            _ingest["progress"] = pct
            _ingest["message"] = msg


# Veto epistémico del agente, en DOS capas:
#  1) Hard floor: si la mejor coincidencia cae por debajo de esto, se abstiene sin llamar al
#     LLM. Conservador a propósito (los embeddings tienen "piso" alto y un umbral agresivo
#     vetaría consultas válidas — hay solapamiento entre off-topic y cubierto).
#  2) Soft: al LLM se le pasa el % de afinidad y se le instruye decir "no me alcanza" si es bajo.
AGENT_VETO_UMBRAL = 0.28

# Techo de videos por playlist. NO es un capricho: cada video = 1 llamada al LLM, y
# Gemini gratis limita req/min. 50 cubre casi cualquier playlist real; subilo si querés
# (a costa de tiempo de ingesta y posibles 429 con listas enormes).
PLAYLIST_LIMIT = 50


def _es_playlist_youtube(url: str) -> bool:
    """URL de playlist 'pura' (no un video suelto que casualmente está en una lista)."""
    u = (url or "").lower()
    if "youtube.com" not in u and "youtu.be" not in u:
        return False
    return "list=" in u and "watch?v=" not in u


def _ingest_playlist(url: str, skip_umap: bool = False):
    """Expande una playlist de YouTube y crea UN NODO POR VIDEO (cada uno enlazable).
    Secuencial (respeta el rate-limit del LLM) y UMAP una sola vez al final."""
    global _ingest
    from processor import expandir_playlist, procesar_youtube
    _set_progress(8, "Leyendo la lista de YouTube…")
    # Listamos TODOS los videos (metadata barata) para saber el total real,
    # pero solo ingerimos los primeros PLAYLIST_LIMIT (cada uno = 1 llamada al LLM).
    videos_all = expandir_playlist(url, limite=200)
    total = len(videos_all)
    if total == 0:
        with _ingest_lock:
            _ingest = {"state": "error",
                       "message": "No se pudieron leer videos de la playlist (¿es pública?)",
                       "label": "", "progress": 0}
        return
    a_cargar = videos_all[:PLAYLIST_LIMIT]
    n = len(a_cargar)
    ok = 0
    for i, v in enumerate(a_cargar):
        titulo = (v.get("title") or "")[:45]
        _set_progress(int(10 + 78 * i / max(1, n)), f"Video {i+1}/{n}: {titulo}…")
        try:
            # Pasamos título y canal de yt-dlp como respaldo (el oembed a veces viene vacío).
            resultado = procesar_youtube(v["url"], title_hint=v.get("title"), author_hint=v.get("channel"))
            for nodo in resultado.get("nodos", []):
                _save_node_sync(nodo)
                ok += 1
        except Exception as e:
            print(f"Error en video {v['url']}: {e}")
    if not skip_umap:
        _set_progress(90, "Calculando posiciones 3D (UMAP)…")
        try:
            import embeddings_engine
            embeddings_engine.main()
        except Exception as e:
            print(f"Error UMAP playlist: {e}")
    if total > PLAYLIST_LIMIT:
        msg = f"Cargué {ok} de {total} videos (tope {PLAYLIST_LIMIT}). Decime si querés el resto."
    else:
        msg = f"{ok} de {total} videos de la playlist incorporados."
    with _ingest_lock:
        _ingest = {"state": "done", "message": msg, "label": "Playlist", "progress": 100}


def _run_ingest(entrada: str, skip_umap: bool = False):
    global _ingest
    try:
        # Playlist de YouTube → un nodo por video (camino propio, sale temprano).
        if entrada.startswith("http") and _es_playlist_youtube(entrada):
            _ingest_playlist(entrada, skip_umap)
            return
        from processor import (
            acumular_resultado,
            procesar_excel,
            procesar_html,
            procesar_pdf,
            procesar_youtube,
        )

        _set_progress(10, "Extrayendo contenido…")

        import threading as _threading
        _stop_ticker = _threading.Event()

        def _ticker():
            steps = [
                (14, "Analizando con IA…"),
                (20, "Analizando con IA…"),
                (27, "Procesando con IA…"),
                (33, "Procesando con IA…"),
                (39, "Generando resumen…"),
                (44, "Generando resumen…"),
                (49, "Extrayendo conceptos…"),
                (54, "Extrayendo conceptos…"),
            ]
            for pct, msg in steps:
                if _stop_ticker.wait(timeout=5):
                    break
                _set_progress(pct, msg)

        _t = _threading.Thread(target=_ticker, daemon=True)
        _t.start()

        ext = Path(entrada).suffix.lower() if not entrada.startswith("http") else ""
        if entrada.startswith("http"):
            yt_domains = ("youtube.com", "youtu.be")
            if any(d in entrada for d in yt_domains):
                resultado = procesar_youtube(entrada)
            else:
                from processor import procesar_url_web
                resultado = procesar_url_web(entrada)
        elif ext in (".xlsx", ".xls"):
            resultado = procesar_excel(entrada)
        elif ext in (".html", ".htm"):
            resultado = procesar_html(entrada)
        elif ext in (".txt", ".md"):
            from processor import procesar_txt
            resultado = procesar_txt(entrada)
        elif ext == ".docx":
            from processor import procesar_word
            resultado = procesar_word(entrada)
        elif ext in (".pptx", ".pptm"):
            from processor import procesar_pptx
            resultado = procesar_pptx(entrada)
        else:
            resultado = procesar_pdf(entrada)

        _stop_ticker.set()
        _set_progress(60, "Generando embeddings…")

        # Acumular solo para contexto de rich_html (fusiona con JSON si existe)
        acumulado = acumular_resultado(resultado)

        _set_progress(75, "Generando apunte IA…")
        try:
            from processor import generar_rich_html
            for new_n in resultado["nodos"]:
                rich_html = generar_rich_html(new_n, acumulado["nodos"], acumulado["relaciones"])
                new_n["rich_html"] = rich_html
                for n in acumulado["nodos"]:
                    if n["id"] == new_n["id"]:
                        n["rich_html"] = rich_html
        except Exception as e:
            print(f"Error generando rich_html en ingesta: {e}")

        # Guardar cada nodo en PostgreSQL
        for nodo in resultado.get("nodos", []):
            _save_node_sync(nodo)

        if not skip_umap:
            _set_progress(85, "Calculando posición 3D (UMAP)…")
            try:
                import embeddings_engine
                embeddings_engine.main()
            except Exception as e:
                print(f"Error al calcular posiciones 3D (UMAP/HDBSCAN): {e}")

        label = resultado["nodos"][0]["label"] if resultado["nodos"] else "Nodo"
        msg = f'"{label}" guardado.' if skip_umap else f'"{label}" incorporado al grafo.'
        with _ingest_lock:
            _ingest = {"state": "done", "message": msg, "label": label, "progress": 100}

    except Exception as exc:
        with _ingest_lock:
            _ingest = {"state": "error", "message": str(exc), "label": "", "progress": 0}


@app.post("/api/ingest")
async def ingest(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(default=None),
    url: str = Form(default=None),
    skip_umap: bool = Form(default=False),
):
    with _ingest_lock:
        if _ingest["state"] == "processing":
            raise HTTPException(409, "Ya hay una ingesta en progreso")
        _ingest.update({"state": "processing", "message": "Iniciando…", "label": "", "progress": 5})

    entradas = []
    if file and file.filename:
        safe_name = Path(file.filename).name
        if not safe_name or safe_name.startswith("."):
            raise HTTPException(400, "Nombre de archivo inválido")
        save_path = UPLOADS / safe_name
        content = await file.read()
        save_path.write_bytes(content)
        entradas.append(str(save_path))
    elif url and url.strip():
        raw_urls = re.split(r"[,\n]+", url)
        entradas = [u.strip() for u in raw_urls if u.strip().startswith("http")]

    if not entradas:
        with _ingest_lock:
            _ingest.update({"state": "idle", "message": "", "label": ""})
        raise HTTPException(400, "Enviá un archivo o al menos una URL válida (http/https)")

    for idx, entrada in enumerate(entradas):
        is_last = idx == len(entradas) - 1
        background_tasks.add_task(_run_ingest, entrada, skip_umap if is_last else True)

    return {"ok": True, "count": len(entradas)}


@app.post("/api/umap-refresh")
async def umap_refresh(background_tasks: BackgroundTasks):
    def _run():
        try:
            import embeddings_engine
            embeddings_engine.main()
        except Exception as e:
            print(f"Error UMAP refresh: {e}")

    background_tasks.add_task(_run)
    return {"ok": True}


@app.post("/api/ingest/reset")
def ingest_reset():
    global _ingest
    with _ingest_lock:
        _ingest = {"state": "idle", "message": "", "label": "", "progress": 0}
    return {"ok": True}


@app.get("/api/ingest/status")
def get_ingest_status():
    with _ingest_lock:
        return dict(_ingest)


# ── Issue module ──────────────────────────────────────────────────────

def _run_issue(descripcion: str, ref_url: str = None, filepath: str = None):
    global _issue_state
    try:
        from processor import procesar_issue, query_llm, _extraer_texto_html, extraer_texto_pdf

        if ref_url:
            with _issue_lock:
                _issue_state.update({"progress": 15, "message": "Obteniendo contenido de la URL…"})
            try:
                import httpx
                headers = {"User-Agent": "Mozilla/5.0 (compatible; Algedi/1.0)"}
                with httpx.Client(timeout=20.0, follow_redirects=True) as c:
                    resp = c.get(ref_url, headers=headers)
                    url_text = _extraer_texto_html(resp.text)[:3000]
                descripcion = f"{descripcion}\n\nContenido de referencia ({ref_url}):\n{url_text}"
            except Exception as e:
                print(f"Warning: URL fetch failed: {e}")

        if filepath:
            with _issue_lock:
                _issue_state.update({"progress": 15, "message": "Extrayendo contenido del archivo…"})
            try:
                ext = Path(filepath).suffix.lower()
                if ext == ".pdf":
                    file_text = extraer_texto_pdf(filepath)[:3000]
                elif ext in (".html", ".htm"):
                    raw = Path(filepath).read_text(encoding="utf-8", errors="ignore")
                    file_text = _extraer_texto_html(raw)[:3000]
                else:
                    file_text = Path(filepath).read_text(encoding="utf-8", errors="ignore")[:3000]
                descripcion = f"{descripcion}\n\nDocumento adjunto ({Path(filepath).name}):\n{file_text}"
            except Exception as e:
                print(f"Warning: file extraction failed: {e}")

        with _issue_lock:
            _issue_state.update({"state": "processing", "message": "Analizando el problema…", "progress": 20, "result": None})

        nodo = procesar_issue(descripcion)

        with _issue_lock:
            _issue_state.update({"progress": 50, "message": "Guardando en el grafo…"})

        _save_node_sync(nodo)

        with _issue_lock:
            _issue_state.update({"progress": 70, "message": "Buscando conexiones relevantes…"})

        # Búsqueda semántica síncrona desde thread
        from processor import get_embed_model, calcular_similitud_coseno
        from database.models import Node as NodeModel
        related_ids = []
        with get_sync_session() as s:
            all_nodes = s.query(NodeModel).filter(
                NodeModel.is_centroid == False,
                NodeModel.embedding.isnot(None),
            ).all()
            try:
                model = get_embed_model()
                vec = model.encode([descripcion], show_progress_bar=False)[0].tolist()
                scored = []
                for n in all_nodes:
                    emb = n.embedding
                    if hasattr(emb, "tolist"):
                        emb = emb.tolist()
                    sim = calcular_similitud_coseno(vec, emb)
                    if sim and sim > 0:
                        scored.append((sim, n.id))
                scored.sort(reverse=True)
                related_ids = [nid for _, nid in scored[:6]]
            except Exception:
                related_ids = []

        related_nodes = []
        with get_sync_session() as s:
            for rid in related_ids:
                n = s.query(NodeModel).filter(
                    NodeModel.id == rid,
                    NodeModel.is_issue == False,
                ).first()
                if n:
                    related_nodes.append(node_to_dict(n))

        context_str = "\n".join([
            f"- **{n['label']}** ({(n.get('fuente') or '').upper()}): {n.get('desc','')}"
            for n in related_nodes
        ]) if related_nodes else "No se encontraron nodos relacionados."

        base_prompt = (
            f"Problema analizado:\n\"{descripcion}\"\n\n"
            f"Conocimiento relevante en el grafo:\n{context_str}\n\n"
        )

        with _issue_lock:
            _issue_state.update({"progress": 30, "message": "Agente 1/4 — Análisis de procesos…"})

        resp_proceso = query_llm(
            [{"role": "user", "content": base_prompt + (
                "Analizá el problema desde la perspectiva de procesos: ¿dónde está la fricción exacta? "
                "¿qué pasos son prescindibles o automatizables? ¿qué métodos del grafo son directamente "
                "aplicables? Respondé en español, 3-4 párrafos."
            )}],
            system=(
                "Sos un experto en análisis y optimización de procesos. Tu foco es identificar "
                "ineficiencias, cuellos de botella y oportunidades de automatización. Sé específico y práctico."
            ),
        )

        with _issue_lock:
            _issue_state.update({"progress": 50, "message": "Agente 2/4 — Gestión de riesgos…"})

        resp_riesgos = query_llm(
            [{"role": "user", "content": base_prompt + (
                "Identificá los riesgos concretos de este problema: ¿qué puede salir mal? "
                "¿qué dependencias son frágiles? ¿cuáles son los puntos de falla críticos? "
                "Listá cada riesgo con probabilidad e impacto estimados. Respondé en español."
            )}],
            system=(
                "Sos especialista en gestión de riesgos y análisis de fallas. Tu foco es identificar "
                "qué puede salir mal, dependencias frágiles, riesgos ocultos y puntos de falla críticos."
            ),
        )

        with _issue_lock:
            _issue_state.update({"progress": 68, "message": "Agente 3/4 — Perspectiva creativa…"})

        resp_creativo = query_llm(
            [{"role": "user", "content": base_prompt + (
                "Proponé 2-3 soluciones creativas e implementables que crucen el conocimiento del grafo "
                "de formas no obvias. Priorizá lo accionable sobre lo teórico. Respondé en español."
            )}],
            system=(
                "Sos un consultor de innovación. Tu foco es proponer soluciones disruptivas y enfoques "
                "no convencionales. Cruzás conocimiento de distintas fuentes para encontrar lo que otros no ven."
            ),
        )

        with _issue_lock:
            _issue_state.update({"progress": 84, "message": "Agente 4/4 — Red Team epistémico…"})

        resp_red_team = query_llm(
            [{"role": "user", "content": base_prompt + (
                "Hacé Red Teaming de este problema. No lo resuelvas — desafiá el marco mental desde "
                "el que se está planteando. Respondé en español, 3-5 puntos concisos."
            )}],
            system=(
                "Tu rol es hacer Red Teaming epistémico. NO respondas la pregunta ni propongas soluciones. "
                "Tu única función es desafiar los supuestos. Identificá: ¿qué da por sentado quien pregunta? "
                "¿Qué sesgo cognitivo podría estar operando (confirmación, anclaje, disponibilidad, "
                "Dunning-Kruger)? ¿Qué no está siendo considerado? ¿Qué consecuencia de segundo orden "
                "está siendo ignorada? Generá fricción cognitiva intencional. Sé directo y sin rodeos."
            ),
        )

        with _issue_lock:
            _issue_state.update({"progress": 95, "message": "Finalizando…"})

        # Persistir el reporte en el nodo: antes vivía solo en este estado transitorio
        # y se perdía al navegar al detalle del issue (quedaba un chat vacío).
        synthesis = {
            "proceso":  resp_proceso,
            "riesgos":  resp_riesgos,
            "creativo": resp_creativo,
            "red_team": resp_red_team,
        }
        try:
            with get_sync_session() as s:
                s.query(NodeModel).filter(NodeModel.id == nodo["id"]).update({"synthesis": synthesis})
                s.commit()
        except Exception as e:
            print(f"Warning: no se pudo persistir el reporte del issue: {e}")

        with _issue_lock:
            _issue_state.update({
                "state": "done",
                "message": f'Issue "{nodo["label"]}" analizado.',
                "progress": 100,
                "result": {
                    "nodo_id": nodo["id"],
                    "label": nodo["label"],
                    "desc": nodo.get("desc", ""),
                    "related_nodes": [
                        {"id": n["id"], "label": n["label"], "desc": n.get("desc", ""), "fuente": n.get("fuente", "")}
                        for n in related_nodes
                    ],
                    "synthesis": synthesis,
                },
            })

    except Exception as exc:
        with _issue_lock:
            _issue_state.update({"state": "error", "message": str(exc), "progress": 0, "result": None})


@app.post("/api/issue")
async def create_issue(
    background_tasks: BackgroundTasks,
    descripcion: str = Form(...),
    url: str = Form(default=None),
    file: UploadFile = File(default=None),
):
    descripcion = descripcion.strip()
    if not descripcion:
        raise HTTPException(400, "Descripción vacía")
    with _issue_lock:
        if _issue_state["state"] == "processing":
            raise HTTPException(409, "Ya hay un issue en proceso")
        _issue_state.update({"state": "processing", "message": "Iniciando…", "progress": 10, "result": None})

    filepath = None
    if file and file.filename:
        safe_name = Path(file.filename).name
        save_path = UPLOADS / safe_name
        content = await file.read()
        save_path.write_bytes(content)
        filepath = str(save_path)

    ref_url = url.strip() if url and url.strip() else None
    background_tasks.add_task(_run_issue, descripcion, ref_url, filepath)
    return {"ok": True}


@app.get("/api/issue/status")
def get_issue_status():
    with _issue_lock:
        return dict(_issue_state)


@app.post("/api/issue/reset")
def reset_issue():
    global _issue_state
    with _issue_lock:
        _issue_state = {"state": "idle", "message": "", "progress": 0, "result": None}
    return {"ok": True}


# ── Static frontend (production) ──────────────────────────────────────
_dist = BASE / "frontend" / "dist"
if _dist.exists():
    app.mount("/", StaticFiles(directory=str(_dist), html=True), name="static")


if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)
